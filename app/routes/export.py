"""Export routes blueprint - PDF and Excel export"""
from flask import Blueprint, render_template, request, send_file, jsonify
from app.models import Section, Timetable, TimeSlot, Faculty, Room, Course
from app import db
from io import BytesIO
import json

export_bp = Blueprint('export', __name__, url_prefix='/export')


@export_bp.route('/')
def index():
    """Export options page"""
    sections = Section.query.filter_by(is_active=True).order_by(Section.semester, Section.name).all()
    faculty = Faculty.query.filter_by(is_active=True).order_by(Faculty.name).all()
    # Room model uses `is_available` instead of `is_active`
    rooms = Room.query.filter_by(is_available=True).order_by(Room.name).all()
    
    return render_template('export/index.html', sections=sections, faculty=faculty, rooms=rooms)


@export_bp.route('/section/<int:section_id>/pdf')
def export_section_pdf(section_id):
    """Export section timetable as PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER
    
    section = Section.query.get_or_404(section_id)
    entries = Timetable.query.filter_by(section_id=section_id).all()
    # Order by day_index (0-4) then period (1-8)
    timeslots = TimeSlot.query.order_by(TimeSlot.day_index, TimeSlot.period).all()
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        alignment=TA_CENTER,
        spaceAfter=20
    )
    elements.append(Paragraph(f"Timetable - Semester {section.semester} Section {section.name}", title_style))
    # Section model stores branch (e.g., CSE) instead of department field
    elements.append(Paragraph(f"Department: {section.branch}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Build timetable grid
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    periods = list(range(1, 9))
    
    # Create entry lookup
    entry_map = {}
    for entry in entries:
        key = (entry.timeslot.day, entry.timeslot.period)
        if key not in entry_map:
            entry_map[key] = []
        entry_map[key].append(entry)
    
    # Header row
    header = ['Day/Period'] + [f'P{p}' for p in periods]
    table_data = [header]
    
    # Time row
    time_row = ['Time']
    slot_times = {s.period: f"{s.start_time}-{s.end_time}" for s in timeslots if s.day == 'Monday'}
    for p in periods:
        time_row.append(slot_times.get(p, ''))
    table_data.append(time_row)
    
    # Data rows
    for day in days:
        row = [day]
        for period in periods:
            entries_at_slot = entry_map.get((day, period), [])
            if entries_at_slot:
                cell_text = []
                for e in entries_at_slot:
                    course_code = e.faculty_course.course.code if e.faculty_course and e.faculty_course.course else ''
                    faculty_code = e.faculty_course.faculty.faculty_id if e.faculty_course and e.faculty_course.faculty else ''
                    room_code = e.room.room_id if e.room else ''
                    batch_text = f" ({e.batch.name})" if e.batch else ""
                    cell_text.append(f"{course_code}{batch_text}\n{faculty_code}\n{room_code}")
                row.append('\n---\n'.join(cell_text))
            else:
                row.append('')
        table_data.append(row)
    
    # Create table
    col_widths = [60] + [75] * 8
    table = Table(table_data, colWidths=col_widths)
    
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#e0e7ff')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 2), (0, -1), colors.HexColor('#f1f5f9')),
        ('FONTNAME', (0, 2), (0, -1), 'Helvetica-Bold'),
    ]))
    
    # Add alternating row colors
    for i in range(2, len(table_data)):
        if i % 2 == 0:
            table.setStyle(TableStyle([
                ('BACKGROUND', (1, i), (-1, i), colors.HexColor('#f8fafc'))
            ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'timetable_sem{section.semester}_{section.name}.pdf'
    )


@export_bp.route('/section/<int:section_id>/excel')
def export_section_excel(section_id):
    """Export section timetable as Excel"""
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    section = Section.query.get_or_404(section_id)
    entries = Timetable.query.filter_by(section_id=section_id).all()
    # Order by day_index (0-4) then period (1-8)
    timeslots = TimeSlot.query.order_by(TimeSlot.day_index, TimeSlot.period).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Sem{section.semester}_{section.name}"
    
    # Styles
    header_fill = PatternFill(start_color="6366f1", end_color="6366f1", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:I1')
    ws['A1'] = f"Timetable - Semester {section.semester} Section {section.name}"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = center_align
    
    # Build data
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    periods = list(range(1, 9))
    
    # Create entry lookup
    entry_map = {}
    for entry in entries:
        key = (entry.timeslot.day, entry.timeslot.period)
        if key not in entry_map:
            entry_map[key] = []
        entry_map[key].append(entry)
    
    # Header row (row 3)
    headers = ['Day'] + [f'Period {p}' for p in periods]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    
    # Time row (row 4)
    ws.cell(row=4, column=1, value='Time').border = thin_border
    slot_times = {s.period: f"{s.start_time}-{s.end_time}" for s in timeslots if s.day == 'Monday'}
    for col, period in enumerate(periods, 2):
        cell = ws.cell(row=4, column=col, value=slot_times.get(period, ''))
        cell.alignment = center_align
        cell.border = thin_border
    
    # Data rows
    for row_idx, day in enumerate(days, 5):
        ws.cell(row=row_idx, column=1, value=day).border = thin_border
        ws.cell(row=row_idx, column=1).font = Font(bold=True)
        
        for col_idx, period in enumerate(periods, 2):
            entries_at_slot = entry_map.get((day, period), [])
            if entries_at_slot:
                cell_text = []
                for e in entries_at_slot:
                    course_code = e.faculty_course.course.code if e.faculty_course and e.faculty_course.course else ''
                    faculty_code = e.faculty_course.faculty.faculty_id if e.faculty_course and e.faculty_course.faculty else ''
                    room_code = e.room.room_id if e.room else ''
                    batch_text = f" ({e.batch.name})" if e.batch else ""
                    cell_text.append(f"{course_code}{batch_text}\n{faculty_code}\n{room_code}")
                cell = ws.cell(row=row_idx, column=col_idx, value='\n'.join(cell_text))
            else:
                cell = ws.cell(row=row_idx, column=col_idx, value='')
            
            cell.alignment = center_align
            cell.border = thin_border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    for col in range(2, 10):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    # Adjust row heights
    for row in range(5, 10):
        ws.row_dimensions[row].height = 60
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'timetable_sem{section.semester}_{section.name}.xlsx'
    )


@export_bp.route('/faculty/<int:faculty_id>/pdf')
def export_faculty_pdf(faculty_id):
    """Export faculty timetable as PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER
    
    faculty = Faculty.query.get_or_404(faculty_id)
    entries = Timetable.query.filter_by(faculty_id=faculty_id).all()
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        alignment=TA_CENTER,
        spaceAfter=20
    )
    elements.append(Paragraph(f"Faculty Timetable - {faculty.name}", title_style))
    elements.append(Paragraph(f"Code: {faculty.faculty_id} | Department: {faculty.department}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Build similar grid as section timetable...
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    periods = list(range(1, 9))
    
    entry_map = {}
    for entry in entries:
        key = (entry.timeslot.day, entry.timeslot.period)
        if key not in entry_map:
            entry_map[key] = []
        entry_map[key].append(entry)
    
    header = ['Day/Period'] + [f'P{p}' for p in periods]
    table_data = [header]
    
    for day in days:
        row = [day]
        for period in periods:
            entries_at_slot = entry_map.get((day, period), [])
            if entries_at_slot:
                cell_text = []
                for e in entries_at_slot:
                    course_code = e.faculty_course.course.code if e.faculty_course and e.faculty_course.course else ''
                    section_label = f"Sem{e.section.semester}-{e.section.name}" if e.section else ''
                    room_code = e.room.room_id if e.room else ''
                    cell_text.append(f"{course_code}\n{section_label}\n{room_code}")
                row.append('\n'.join(cell_text))
            else:
                row.append('')
        table_data.append(row)
    
    col_widths = [60] + [80] * 8
    table = Table(table_data, colWidths=col_widths)
    
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#f1f5f9')),
    ]))
    
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'timetable_faculty_{faculty.code}.pdf'
    )


@export_bp.route('/room/<int:room_id>/pdf')
def export_room_pdf(room_id):
    """Export room timetable as PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER
    
    room = Room.query.get_or_404(room_id)
    entries = Timetable.query.filter_by(room_id=room_id).all()
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        alignment=TA_CENTER,
        spaceAfter=20
    )
    elements.append(Paragraph(f"Room Timetable - {room.name}", title_style))
    elements.append(Paragraph(f"Code: {room.room_id} | Capacity: {room.capacity}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    periods = list(range(1, 9))
    
    entry_map = {}
    for entry in entries:
        key = (entry.timeslot.day, entry.timeslot.period)
        if key not in entry_map:
            entry_map[key] = []
        entry_map[key].append(entry)
    
    header = ['Day/Period'] + [f'P{p}' for p in periods]
    table_data = [header]
    
    for day in days:
        row = [day]
        for period in periods:
            entries_at_slot = entry_map.get((day, period), [])
            if entries_at_slot:
                cell_text = []
                for e in entries_at_slot:
                    course_code = e.faculty_course.course.code if e.faculty_course and e.faculty_course.course else ''
                    section_label = f"Sem{e.section.semester}-{e.section.name}" if e.section else ''
                    faculty_code = e.faculty_course.faculty.faculty_id if e.faculty_course and e.faculty_course.faculty else ''
                    cell_text.append(f"{course_code}\n{section_label}\n{faculty_code}")
                row.append('\n'.join(cell_text))
            else:
                row.append('')
        table_data.append(row)
    
    col_widths = [60] + [80] * 8
    table = Table(table_data, colWidths=col_widths)
    
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f59e0b')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#f1f5f9')),
    ]))
    
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'timetable_room_{room.code}.pdf'
    )


@export_bp.route('/master/pdf')
def export_master_pdf():
    """Export master timetable (all sections) as PDF"""
    # Implementation for master timetable
    pass


@export_bp.route('/master/excel')
def export_master_excel():
    """Export master timetable as Excel"""
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    sections = Section.query.filter_by(is_active=True).order_by(Section.semester, Section.name).all()
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    for section in sections:
        entries = Timetable.query.filter_by(section_id=section.id).all()
        if not entries:
            continue
        
        ws = wb.create_sheet(title=f"Sem{section.semester}_{section.name}")
        
        # Similar logic as export_section_excel
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        periods = list(range(1, 9))
        
        entry_map = {}
        for entry in entries:
            key = (entry.timeslot.day, entry.timeslot.period)
            if key not in entry_map:
                entry_map[key] = []
            entry_map[key].append(entry)
        
        # Headers
        ws.cell(row=1, column=1, value='Day')
        for col, p in enumerate(periods, 2):
            ws.cell(row=1, column=col, value=f'P{p}')
        
        # Data
        for row_idx, day in enumerate(days, 2):
            ws.cell(row=row_idx, column=1, value=day)
            for col_idx, period in enumerate(periods, 2):
                entries_at_slot = entry_map.get((day, period), [])
                if entries_at_slot:
                    cell_text = []
                    for e in entries_at_slot:
                        course_code = e.faculty_course.course.code if e.faculty_course and e.faculty_course.course else ''
                        batch_text = f" ({e.batch.name})" if e.batch else ""
                        cell_text.append(f"{course_code}{batch_text}")
                    ws.cell(row=row_idx, column=col_idx, value='\n'.join(cell_text))
    
    if not wb.sheetnames:
        # No timetables to export
        ws = wb.create_sheet(title="No Data")
        ws['A1'] = "No timetables generated yet"
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='master_timetable.xlsx'
    )
